from openpyxl import Workbook, load_workbook
from datetime import datetime as dt
import operator, os

def auto_col_width(worksheet):

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            if cell.coordinate in worksheet.merged_cells: # not check merge_cells
                continue
            try: # Necessary to avoid error on empty cells
                if type(cell.value) == dt:
                    if len(dt.date(cell.value).strftime('%x %X')) > max_length:
                        max_length = len(dt.date(cell.value).strftime('%x %X'))
                else:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

def isfloat(value):

    try:
        float(value)
        return True
    except ValueError:
        return False


def dict2xl(myDict, dest_filename, sheettitle=None, colList=None):

    if not colList:
        colList = list(
            myDict[
                max(
                    [(k,len(m)) for k, m in enumerate(myDict)],
                    key=operator.itemgetter(1))[0]
                ].keys() if myDict else []
            )


    myList = [colList] # 1st row = header

    for item in myDict: myList.append([item.get(col,'') for col in colList])

    if os.path.exists(dest_filename):
        wb = load_workbook(dest_filename)
        if sheettitle in wb.sheetnames:
            del wb[sheettitle]
        ws = wb.create_sheet()
    else:
        wb = Workbook()
        ws = wb.active

    if sheettitle: ws.title = sheettitle
    wb.guess_type = True

    for row in range(1,len(myList)):
        ws.append(myList[row-1])

    auto_col_width(ws)
    
    wb.save(filename=dest_filename)
    wb.close()
