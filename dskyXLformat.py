from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl import Workbook, load_workbook
from datetime import date, datetime
import operator, os

valrule = DataBarRule(
    start_type='num',
    start_value=0,
    end_type='max',
    color="FF638EC6",
    showValue="None",
    )

pcntrule = DataBarRule(
    start_type='num',
    start_value=0,
    end_type='num',
    end_value=1,
    color="FF638EC6",
    showValue="None",
    )

humrule = DataBarRule(
    start_type='num',
    start_value=0,
    end_type='num',
    end_value=1,
    color="FFF79646",
    showValue="None",
    )

cloudrule = DataBarRule(
    start_type='num',
    start_value=0,
    end_type='num',
    end_value=1,
    color="FFBFBFBF",
    showValue="None",
    )

temprule = ColorScaleRule(
    start_type='min',
    start_color='5A8AC6',
    mid_type='percentile',
    mid_value=50,
    mid_color='FCFCFF',
    end_type='max',
    end_color='F8696B'
    )

dewrule = ColorScaleRule(
    start_type='min',
    start_color='63BE7B',
    end_type='max',
    end_color='FFEF9C'
    )

windrule = DataBarRule(
    start_type='num',
    start_value=0,
    end_type='max',
    color="FF00CC99",
    showValue="None",
    )

gustrule = DataBarRule(
    start_type='num',
    start_value=0,
    end_type='max',
    color="FF00FFFF",
    showValue="None",
    )

pressrule = ColorScaleRule(
    start_type='min',
    start_color='A6A6A6',
    end_type='max',
    end_color='FFFFFF'
    )

def cond_format(worksheet, column, formattype='value'):

    fmtrule = {
        'value': valrule,
        'pcnt': pcntrule,
        'hum': humrule,
        'cloud': cloudrule,
        'temp': temprule,
        'dew': dewrule,
        'wind': windrule,
        'gust': gustrule,
        'pressure': pressrule
        }

    col_letter = column[0].column
    colstring = col_letter + '2:' + col_letter + str(len(worksheet[col_letter]))
    dbar_rule = fmtrule[formattype]
    
    worksheet.conditional_formatting.add(
        colstring,
        dbar_rule
        )
    
def mint(worksheet):
    # {columnname: [dataformat, conditional_format_type, format_rule]}
    numformat = {
        'time': ['ddd h:mm AM/PM', 'None'],
        'precipIntensity': ['General \i\\n\/\hr', 'value'],
        'precipIntensityError': ['General σ', 'None'],
        'precipProbability': ['0%', 'pcnt'],
        'precipType': ['Text', 'None']
        }
    
    for col in worksheet.columns:
        for cell in col:
            cell.number_format = numformat[col[0].value][0]
        if numformat[col[0].value][1] != 'None':
            cfrule = numformat[col[0].value][1]
            cond_format(worksheet, col, cfrule)
    
def hrly(worksheet):
    numformat = {
        'time': ['ddd mmm dd, h AM/PM', 'None'],
        'summary': ['Text', 'None'],
        'icon': ['Text', 'None'],
        'precipIntensity': ['General \i\\n\/\hr', 'value'],
        'precipIntensityError': ['General σ', 'None'],
        'precipProbability': ['0%', 'pcnt'],
        'precipType': ['Text', 'None'],
        'precipAccumulation': ['General \i\\n', 'value'],
        'temperature': ['General°F', 'temp'],
        'apparentTemperature': ['General°F', 'temp'],
        'dewPoint': ['General°F', 'dew'],
        'humidity': ['0%', 'hum'],
        'pressure': ['General \i\\n\H\g', 'pressure'],
        'windSpeed': ['General \mp\h', 'wind'],
        'windGust': ['General \mp\h', 'gust'],
        'windBearing': ['General°', 'None'],
        'cloudCover': ['0%', 'cloud'],
        'uvIndex': ['General', 'None'],
        'visibility': ['General \m\i', 'value'],
        'ozone': ['General \DU', 'None']
        }
    
    for col in worksheet.columns:
        for cell in col:
            cell.number_format = numformat[col[0].value][0]
        if numformat[col[0].value][1] != 'None':
            cfrule = numformat[col[0].value][1]
            cond_format(worksheet, col, cfrule)

def daze(worksheet):
    dayformat={
        'time': ['ddd mmm dd', 'None'],
        'summary': ['Text', 'None'],
        'icon': ['Text', 'None'],
        'sunriseTime': ['h:mm AM/PM', 'None'],
        'sunsetTime': ['h:mm AM/PM', 'None'],
        'moonPhase': ['General', 'None'],
        'precipIntensity': ['General \i\\n\/\hr', 'value'],
        'precipIntensityMax': ['General \i\\n\/\hr', 'value'],
        'precipIntensityMaxTime': ['h:mm AM/PM', 'None'],
        'precipProbability': ['0%', 'pcnt'],
        'precipAccumulation': ['General \i\\n', 'value'],
        'precipType': ['Text', 'None'],
        'temperatureHigh': ['General°F', 'temp'],
        'temperatureHighTime': ['h:mm AM/PM', 'None'],
        'temperatureLow': ['General°F', 'temp'],
        'temperatureLowTime': ['h:mm AM/PM', 'None'],
        'apparentTemperatureHigh': ['General°F', 'temp'],
        'apparentTemperatureHighTime': ['h:mm AM/PM', 'None'],
        'apparentTemperatureLow': ['General°F', 'temp'],
        'apparentTemperatureLowTime': ['h:mm AM/PM', 'None'],
        'dewPoint': ['General°F', 'dew'],
        'humidity': ['0%', 'hum'],
        'pressure': ['General \i\\n\H\g', 'pressure'],
        'windSpeed': ['General \mp\h', 'wind'],
        'windGust': ['General \mp\h', 'gust'],
        'windGustTime': ['h:mm AM/PM', 'None'],
        'windBearing': ['General°', 'None'],
        'cloudCover': ['0%', 'cloud'],
        'uvIndex': ['General', 'None'],
        'uvIndexTime': ['h:mm AM/PM', 'None'],
        'visibility': ['General \m\i', 'value'],
        'ozone': ['General \DU', 'None'],
        'temperatureMin': ['General°F', 'temp'],
        'temperatureMinTime': ['h:mm AM/PM', 'None'],
        'temperatureMax': ['General°F', 'temp'],
        'temperatureMaxTime': ['h:mm AM/PM', 'None'],
        'apparentTemperatureMin': ['General°F', 'temp'],
        'apparentTemperatureMinTime': ['h:mm AM/PM', 'None'],
        'apparentTemperatureMax': ['General°F', 'temp'],
        'apparentTemperatureMaxTime': ['h:mm AM/PM', 'None']
        }
    
    for col in worksheet.columns:
        for cell in col:
            cell.number_format = dayformat[col[0].value][0]
        if dayformat[col[0].value][1] != 'None':
            cfrule = dayformat[col[0].value][1]
            cond_format(worksheet, col, cfrule)

    
def dsky_format_xlbook(filename):

    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        return False

    for sheet in wb.sheetnames:
        if sheet == 'Minutely':
            mint(wb[sheet])
            wb[sheet].freeze_panes = wb[sheet]['A2']
        elif sheet == 'Hourly':
            hrly(wb[sheet])
            wb[sheet].freeze_panes = wb[sheet]['B2']
        elif sheet == 'Daily':
            daze(wb[sheet])
            wb[sheet].freeze_panes = wb[sheet]['B2']
        else:
            # unknown sheet
            pass

    wb.save(filename=filename)
    wb.close()
    
    return True
